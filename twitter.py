#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
Download the Twitter history for a certain user account

Author: Aaron Dettmann
 _________________________________________
/ Fame is a vapor; popularity an          \
| accident; the only earthly certainty is |
\ oblivion. -- Mark Twain                 /
 -----------------------------------------
        \   ^__^
         \  (oo)\_______
            (__)\       )\/\
                ||----w |
                ||     ||
"""

from collections import Counter, OrderedDict
from functools import partial
from pathlib import Path
import argparse
import datetime
import json
import logging
import os
import sys

# ===== Check non-standard libraries =====
__import_errors__ = ''
try:
    import matplotlib.pyplot as plt
except ImportError:
    __import_errors__ += "Please install 'matplotlib'\n"

try:
    import openpyxl as xl
except ImportError:
    __import_errors__ += "Please install 'openpyxl'\n"

try:
    import twitter_scraper as tw
except ImportError:
    __import_errors__ += "Please install 'twitter_scraper'"

if __import_errors__:
    print(__import_errors__, file=sys.stderr)
    sys.exit(1)

__prog_name__ = 'TwitterHistory'

HERE = os.path.abspath(os.path.dirname(__file__))
DIR_DATA = os.path.join(HERE, 'data')

LOG_FMT = '%(asctime)s | %(levelname)s | %(message)s'
LOG_DATE_FMT = '%F %H:%M:%S'
logging.basicConfig(level=logging.INFO, format=LOG_FMT, datefmt=LOG_DATE_FMT)


# See https://stackoverflow.com/questions/12122007/python-json-encoder-to-support-datetime
class DateTimeEncoder(json.JSONEncoder):
    def default(self, obj):
        if isinstance(obj, (datetime.datetime, datetime.date, datetime.time)):
            return obj.isoformat()
        elif isinstance(obj, datetime.timedelta):
            return (datetime.datetime.min + obj).time().isoformat()

        return super(DateTimeEncoder, self).default(obj)


dump_pretty_json = partial(json.dump, cls=DateTimeEncoder, indent=4, separators=(',', ': '))


def cli():
    parser = argparse.ArgumentParser(prog=f'{__prog_name__}')
    subparsers = parser.add_subparsers(help='execution modes', dest='exec_mode')

    # ----- Mode 'down' -----
    sub = subparsers.add_parser('down', help='download target twitter feed')
    sub.add_argument('usernames', metavar='NAMES', nargs='+', type=str, help='target twitter profile')
    sub.add_argument('--pages', '-p', metavar='PAGES', type=int, help='number of pages to fetch', default=200)
    sub.add_argument('--no-excel', action='store_true', help='do not convert data to exel file')

    # ----- Mode plot -----
    sub = subparsers.add_parser('plot', help='plot twitter feed data')
    sub.add_argument("filename", metavar='FILE', help="data to visualize", type=str)
    sub.add_argument("--show", "-s", metavar='PLOTNAME', help="available plots: 'activity'", type=str, default='activity')

    # ----- Mode 'xl' -----
    sub = subparsers.add_parser('xl', help='convert data to excel spreadsheet')
    sub.add_argument("filename", metavar='FILE', help="data to convert", type=str)

    args = parser.parse_args()

    if args.exec_mode == 'down':
        for username in args.usernames:
            # Remove commas
            username = username.strip().strip(',')

            json_file = download_history(username, args.pages)
            excel_file = json_file.replace('.json', '.xlsx')
            # Convert data to Excel spreadsheet by default
            if not args.no_excel:
                twitter_data = load_twitter_data(json_file)
                convert_to_excel(twitter_data, excel_file)
    elif args.exec_mode == 'plot':
        filename = os.path.abspath(args.filename)
        twitter_data = load_twitter_data(filename)
        plot_tweet_activity(twitter_data)
    elif args.exec_mode == 'xl':
        json_file = os.path.abspath(args.filename)
        excel_file = json_file.replace('.json', '.xlsx')

        twitter_data = load_twitter_data(json_file)
        convert_to_excel(twitter_data, excel_file)
    else:
        parser.print_help()


def mkdir(dirname):
    """
    Make a directory

    Args:
        :dirname: (str) path name
    """

    Path(dirname).mkdir(parents=True, exist_ok=True)


def load_twitter_data(filename):
    """
    Load twitter data from a file

    Args:
        :filename: (str) path of the file to load

    Returns:
        :twitter_data: (dict) dictionary with twitter data
    """

    logging.info(f"Importing twitter data from file: {filename}")
    with open(str(filename), "r") as fp:
        twitter_data = json.load(fp)

    logging.info(f"Found {len(twitter_data['history'])} tweets in imported file...")
    return twitter_data


def download_history(username, pages):
    """
    Download tweets and save data on disk

    Args:
        :username: (str) target twitter account
        :pages: (int) number of pages to download
    """

    try:
        profile = tw.Profile(username)
        logging.info(f"Target: {username} ({profile.name}) | {profile.followers_count:,} followers")
    except:
        if username.startswith('#'):
            logging.info(f"Interpreting {username!r} as a hashtag...")
        else:
            logging.error(f"Failed to fetch username data for {username!r}")
        profile = None

    logging.info(f"Downloading tweets ({pages} pages)...")
    tweets = list(tw.get_tweets(username, pages))
    logging.info(f"Downloaded {len(tweets)} tweets...")

    mkdir(DIR_DATA)
    now = datetime.datetime.now()
    dir_user_data = os.path.join(DIR_DATA, f"{username}_{now.strftime('%F_%H%M')}")
    mkdir(dir_user_data)
    file_user_data = os.path.join(dir_user_data, "data.json")

    profile_dict = profile.to_dict() if profile is not None else {}
    data = {
        "profile": profile_dict,
        "history": tweets,
    }
    logging.info(f"Saving data: {file_user_data}")
    with open(file_user_data, 'w') as fp:
        dump_pretty_json(data, fp)
    return file_user_data


def plot_tweet_activity(twitter_data):
    """
    Visualize daily tweet activity

    Args:
        :twitter_data: (dict) dictionary with twitter data
    """

    tweets = twitter_data['history']
    profile = twitter_data['profile']
    username = profile.get('username', 'NONE')

    tweets_per_day = get_tweets_per_day(twitter_data)

    logging.info(f"Plotting tweet activity for {username}...")
    dates = list(tweets_per_day.keys())
    num_tweets = list(tweets_per_day.values())

    plt.plot_date(dates, num_tweets, '.', color='black')
    plt.gcf().autofmt_xdate()
    plt.title(f"Tweet activity @{username} (total = {len(tweets)})")
    plt.xlabel("Time")
    plt.ylabel("Number of tweets")
    plt.show()


def daterange(start_date, end_date):
    # See https://stackoverflow.com/questions/1060279/iterating-through-a-range-of-dates-in-python
    for n in range(int((end_date - start_date).days)):
        yield start_date + datetime.timedelta(n)


def get_tweets_per_day(twitter_data, count_zero_days=True):
    """
    Return tweets per day

    Args:
        :twitter_data: (dict) dictionary with twitter data
        :count_zero_days: (bool) include day with zero tweets

    Returns:
        :tweets_per_day: (dict) dictionary with day (datetime object) and tweet count
    """

    tweets = get_tweets(twitter_data)

    tweets_per_day = Counter()
    for tweet in tweets:
        time = datetime.datetime.fromisoformat(tweet['time'])
        tweets_per_day[datetime.datetime(time.year, time.month, time.day)] += 1

    # Fill up the dictionary with zeros
    if count_zero_days:
        logging.info("Looking for days with zero tweets...")

        # Get start and end date
        all_dates = sorted(list(tweets_per_day.keys()))  # list of sorted dates
        start_date, end_date = all_dates[0], all_dates[-1]
        if end_date < start_date:
            start_date, end_date, = end_date, start_date

        logging.info(f"Date range is {start_date.strftime('%F')} to {end_date.strftime('%F')}")
        for day in daterange(start_date, end_date):
            tweets_per_day[datetime.datetime(day.year, day.month, day.day)] += 0

    tweets_per_day = _sort_date_dict(tweets_per_day)
    return tweets_per_day


def _sort_date_dict(date_dict):
    """
    Sort a dictionary with keys as dates in chronological order

    Args:
        :date_dict: (dict) dictionary with dates

    Returns:
        :sorted_dict: (dict) dictionary sorted by keys
    """

    # OrderedDict() should not be necessary in 3.8
    sorted_dict = OrderedDict({k: v for k, v in sorted(date_dict.items())})
    return sorted_dict


def get_tweets(twitter_data):
    """
    Return tweets from twitter data or return error

    Args:
        :twitter_data: (dict) dictionary with twitter data

    Returns:
        :tweets: (list) Tweets
    """

    tweets = twitter_data.get('history', None)
    if tweets is None:
        logging.error("Failed to retrieve tweets... Exit.")
        sys.exit(1)
    return tweets


def convert_to_excel(twitter_data, excel_file):
    """
    Convert a twitter data dictionary to a excel file

    Args:
        :twitter_data: (dict) dictionary with twitter data
        :excel_file: (str) excel file name
    """

    logging.info("Creating excel file...")
    tweets = get_tweets(twitter_data)

    workbook = xl.Workbook()

    # ----- Tweet data -----
    sheet1 = workbook.active
    sheet1.title = "Tweet raw data"

    headers = ["Time", "isRetweet", "replies", "likes", "hashtags", "text"]
    for i, header in enumerate(headers, start=1):
        sheet1.cell(row=1, column=i, value=header)

    for i, tweet in enumerate(tweets, start=2):
        sheet1.cell(row=i, column=1, value=tweet['time'])
        sheet1.cell(row=i, column=2, value=tweet['isRetweet'])
        sheet1.cell(row=i, column=3, value=tweet['replies'])
        sheet1.cell(row=i, column=4, value=tweet['likes'])
        sheet1.cell(row=i, column=5, value=str(tweet['entries']['hashtags']))
        sheet1.cell(row=i, column=6, value=tweet['text'])

    # ----- Tweets per day -----
    sheet2 = workbook.create_sheet(title="Twitter activity")
    for i, header in enumerate(["Time", "numTweets"], start=1):
        sheet2.cell(row=1, column=i, value=header)

    tweets_per_day = get_tweets_per_day(twitter_data)
    for i, (day, num_tweets) in enumerate(tweets_per_day.items(), start=2):
        sheet2.cell(row=i, column=1, value=day.strftime('%F'))
        sheet2.cell(row=i, column=2, value=num_tweets)

    # ----- User data -----
    sheet3 = workbook.create_sheet(title="Account")
    profile = twitter_data['profile']
    headers = ["name", "username", "likes_count", "tweets_count", "followers_count", "following_count"]
    for i, header in enumerate(headers, start=1):
        sheet3.cell(row=i, column=1, value=header)
        sheet3.cell(row=i, column=2, value=profile.get(header, 'NONE'))

    logging.info(f"Saving data: {excel_file}")
    workbook.save(excel_file)


if __name__ == '__main__':
    try:
        cli()
    except KeyboardInterrupt:
        logging.error(f"Exit...")
        sys.exit(1)
